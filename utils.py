import smtplib
from openpyxl import Workbook
from email.message import EmailMessage
from openpyxl.styles import PatternFill, Font
from logger import logs
from datetime import datetime


def create_excel_file(filename, da, customer_name, date, address):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = list(da[0].keys()) if da else []

    # Add customer information at the top
    ws['A1'] = f"Customer Name: {customer_name}"
    ws['A2'] = f"Date: {date}"
    ws['A3'] = f"Address: {address}"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30

    # Add headers and style
    ws.append(headers)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows
    for row in da:
        ws.append([row.get(header, "") for header in headers])

    wb.save(filename)
    print(f"{filename} saved successfully.")
    logs.insert_one({"timestamp": datetime.utcnow(), "content": f"{filename} saved successfully."})


def send_email_with_attachment(subject, body, sender_email, sender_password, receiver_email, attachment_path):
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender_email
    msg["To"] = receiver_email
    msg.set_content(body)

    with open(attachment_path, "rb") as f:
        file_data = f.read()
        msg.add_attachment(file_data, maintype="application",
                           subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           filename=attachment_path)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(sender_email, sender_password)
        smtp.send_message(msg)

    print("Email sent successfully.")
    logs.insert_one({"timestamp": datetime.utcnow(),"content": f"Email sent successfully."})
